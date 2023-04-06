from bs4 import BeautifulSoup

from datetime import date
import selenium
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import time
from openpyxl import load_workbook
import xlsxwriter


def scrapePage(driver):
    products = {}
    driver.execute_script("window.scrollBy(0,document.body.scrollHeight)")
    productTiles = driver.find_elements(By.CLASS_NAME,"product-tile-v2")
    for product in productTiles:
        title = product.find_element(By.CLASS_NAME,"product-title-link").text
        try:
            price = product.find_element(By.CLASS_NAME,"product-tile-price")
            pricenumber = price.text.split('\n')[0].replace("$","")
        except:
            pricenumber = -1
        products[title] = pricenumber
        print(f"{title}: {pricenumber}")
    return products


def nextPage(driver):
    button = driver.find_element(By.CLASS_NAME,"paging-next")
    button.click()

    time.sleep(10)
    driver.execute_script("""
    var element = document.querySelector(".container-carousel");
    if (element)
        element.parentNode.removeChild(element);
    """)
    


def saveData(products, page):
    print("\n\nsaving...")
    row = (today - start_date).days + 2
    workbook = load_workbook("data.xlsx")
    sheet = workbook[page]
    #products.update({"EPIC SECRET":"99999","NAH HOW":"888"})
    sheet.cell(row = row, column = 1).value = today
    productnames = list(products.keys())
    productnames.sort()
    productCol = 0
    skipped = 0
    toInsert = []
    for name in productnames:
        print(f"{name}: {products[name]}")
    for i in range(len(products)):
        productname = productnames[i]
        productprice = products[productnames[i]]
        print(f"i {i}")
        cell = sheet.cell(row=1, column = i +2+skipped)
        print(f"current product  {productname}: {productprice}, cell column: {cell.value}")
        if(productname == cell.value):
            print("foundnormal")
            sheet.cell(row = row,column = i +2+skipped).value = productprice
        elif(cell.value == None):
            print("null")
            cell.value = productname
            sheet.cell(row = row,column = i +2+skipped).value = productprice
        else:
            checker = [str(sheet.cell(row = 1,column = i +2+skipped).value),productname]
            while(checker[0]!=productname):
                checker.sort()
                #a = input("epic debug")
                print(f" check 0{checker[0]}, check 1 {checker[1]}, product name {productname}, i,  {i}")
                if(checker[0] == checker[1]):
                    print("found location (not possible in test)")
                    sheet.cell(row = row,column = i +2+skipped).value = productprice
                    break
                
                if(checker[0]==productname):# check b, at c found error insert at c location
                    # a|c
                    print("ideal found")
                    sheet.insert_cols(i+  +2+skipped)
                    cell = sheet.cell(row=1, column = i +2+skipped)
                    cell.value = productname
                    sheet.cell(row = row,column = i +2+skipped).value = productprice
                    break
                if(sheet.cell(row = 1,column = i +2+skipped).value == None):
                    cell = sheet.cell(row=1, column = i +2+skipped)
                    cell.value = productname
                    sheet.cell(row = row,column = i +2+skipped).value = productprice
                    break
                checker[1] = str(sheet.cell(row = 1,column = i +2+skipped).value)
                checker[0] = productname
                checker.sort()
                skipped+=1
                print(f"checker {checker} val of column {str(sheet.cell(row = 1,column = i +2+skipped).value)}")
                
    workbook.save(r'data.xlsx')
    print("Finished Saving")
def Insert(products):
    print("real code")

def scrapeAllPage(URL):
    driver = webdriver.Firefox()
    driver.get(URL)

    time.sleep(5)
    driver.execute_script("""
    var element = document.querySelector(".container-carousel");
    if (element)
        element.parentNode.removeChild(element);
    """)
    time.sleep(5)
    try:
        pageselector = driver.find_element(By.CLASS_NAME,"paging-section")
        print(f"pageselector: {pageselector.text}")
        pagecounts = pageselector.find_elements(By.CLASS_NAME,'paging-pageNumber')
        pagecount = pagecounts[-1].text
        print(f"page count: {pagecount}")
    except Exception as e:
            print(str(e.__class__))
            print("so it failed L")
    
    a = input("degub 101")



    for i in range(int(pagecount)): 
        print(f"i is {i}")
        try:
            products.update(scrapePage(driver))
        except Exception as e:
            print(str(e.__class__))
            driver.refresh()
            print("Element not found refreshing... \n\n\n")
            time.sleep(10)
            try:
                products.update(scrapePage(driver))
            except:
                break
        driver.get(f"https://www.woolworths.com.au/shop/browse/fruit-veg?pageNumber={i+2}")
        time.sleep(5)
        try:
            print("legit next")
        except:
            print("Error clicking next page")
            driver.refresh()
            time.sleep(10)
            try:
                driver.get(f"https://www.woolworths.com.au/shop/browse/fruit-veg?pageNumber={i+2}")
            except:
                print("two times filed assuming end of pages")
                break
    saveData(products,URL.split("/")[-1])
    


URLS =["https://www.woolworths.com.au/shop/browse/fruit-veg"]
""",
        "https://www.woolworths.com.au/shop/browse/snacks-confectionery",
        "https://www.woolworths.com.au/shop/browse/lunch-box",
        "https://www.woolworths.com.au/shop/browse/freezer",
        "https://www.woolworths.com.au/shop/browse/drinks",
        "https://www.woolworths.com.au/shop/browse/meat-seafood-deli",
        "https://www.woolworths.com.au/shop/browse/bakery",
        "https://www.woolworths.com.au/shop/browse/dairy-eggs-fridge"]
"""
""" 
        fucking massive
        "https://www.woolworths.com.au/shop/browse/pantry",

    who cares    
        "https://www.woolworths.com.au/shop/browse/health-wellness",

        "https://www.woolworths.com.au/shop/browse/liquor",
        "https://www.woolworths.com.au/shop/browse/pet",
        "https://www.woolworths.com.au/shop/browse/baby",
        "https://www.woolworths.com.au/shop/browse/beauty-personal-care",
        "https://www.woolworths.com.au/shop/browse/household"

"""

wb = load_workbook("data.xlsx")
for i in URLS:
    wb.create_sheet(i.split("/")[-1])
wb.save("data.xlsx")
start_date = date.fromisocalendar(2023, 12, 4)
print(start_date)
today = date.today()
print((today - start_date).days + 2)

for pageName in URLS:
    products = {}
    scrapeAllPage(pageName)

